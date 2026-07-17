"""Live dashboard + control panel for the config-driven scraper.

The DB is the source of truth for configs, so this app both reads (counts,
runs, fill rates, records) and writes (create/edit/delete a site, probe it,
launch runs). Runs execute in background threads and stream records page by
page, so the table and counts fill in live.

Run (host):   POSTGRES_HOST=localhost uvicorn dashboard:app --port 8050
Run (docker): docker compose up dashboard
Then open     http://localhost:8050
"""
import csv
import datetime
import io
import os
import secrets
import tempfile
import threading

from fastapi import Body, Depends, FastAPI, HTTPException
from fastapi.responses import FileResponse, HTMLResponse, StreamingResponse
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from starlette.background import BackgroundTask
from openpyxl import Workbook
from openpyxl.styles import Font

import config
import db
import runner
import schema
import scripts_api
import sites_api

# HTTP Basic auth on every route in this app (dashboard + all included
# routers below). The dashboard is tunneled to the public internet, so an
# unset DASHBOARD_PASSWORD must deny everything rather than allow it.
_basic = HTTPBasic()


def _require_auth(creds: HTTPBasicCredentials = Depends(_basic)) -> None:
    if not config.DASHBOARD_PASSWORD:
        raise HTTPException(
            status_code=503,
            detail="DASHBOARD_PASSWORD is not set; dashboard is locked until it is configured.",
        )
    user_ok = secrets.compare_digest(creds.username, config.DASHBOARD_USER)
    pass_ok = secrets.compare_digest(creds.password, config.DASHBOARD_PASSWORD)
    if not (user_ok and pass_ok):
        raise HTTPException(
            status_code=401,
            detail="Invalid credentials",
            headers={"WWW-Authenticate": "Basic"},
        )


app = FastAPI(title="Scraper monitor", dependencies=[Depends(_require_auth)])
app.include_router(scripts_api.router)
import dedupe_api
app.include_router(dedupe_api.router)
app.include_router(sites_api.router)

@app.on_event("startup")
def _cleanup_orphans():
    # A fresh process can't have any live runs, so mark leftover 'running'
    # rows (orphaned by the previous restart/crash) as interrupted.
    try:
        db.init()
        db.mark_orphan_runs()
    except Exception:
        pass

# in-process guard so the same site can't run twice at once
_running = set()
_cancel = set()
_lock = threading.Lock()


def _start_bg(cfg) -> bool:
    with _lock:
        if cfg.name in _running:
            return False
        _running.add(cfg.name)
        _cancel.discard(cfg.name)

    def work():
        try:
            runner.execute_run(cfg, should_stop=lambda: cfg.name in _cancel)
        except Exception:
            pass
        finally:
            with _lock:
                _running.discard(cfg.name)
                _cancel.discard(cfg.name)

    threading.Thread(target=work, daemon=True).start()
    return True


# ---- reads ---------------------------------------------------------------
@app.get("/api/sites")
def api_sites():
    return {"sites": db.sites_overview(), "running": sorted(_running)}


@app.get("/api/runs")
def api_runs(limit: int = 20):
    return {"runs": db.recent_runs(limit)}


@app.delete("/api/runs")
def api_clear_runs():
    return {"ok": True, "deleted": db.delete_runs()}


@app.get("/api/stats")
def api_stats(site: str):
    return {"site": site, "fields": db.field_fill_latest(site)}


@app.get("/api/preview")
def api_preview(site: str, limit: int = 60, q: str = ""):
    fields = db.site_fields(site)
    rows = [{"_rid": r["id"], **(r["data"] or {})} for r in db.preview(site, limit, q)]
    if not fields and rows:
        seen = []
        for r in rows:
            for k in r:
                if k not in seen and not k.startswith("_"):
                    seen.append(k)
        fields = seen
    return {"fields": fields, "rows": rows}


@app.delete("/api/records")
def api_delete_records(site: str = "", ids: str = "", all: bool = False):
    id_list = [int(x) for x in ids.split(",") if x.strip().lstrip("-").isdigit()] \
        if ids else None
    n = db.delete_records(site=site or None, ids=id_list, all_sites=all)
    return {"ok": True, "deleted": n}


@app.get("/api/site")
def api_get_site(name: str):
    return {"config": db.get_site(name)}


# ---- writes --------------------------------------------------------------
@app.post("/api/site")
def api_save_site(payload: dict = Body(...)):
    try:
        cfg = schema.from_dict(payload)
    except Exception as e:
        return {"ok": False, "error": str(e)}
    db.init()
    db.upsert_site(cfg.name, cfg.url, cfg.engine, cfg.raw)
    return {"ok": True, "name": cfg.name}


@app.delete("/api/site")
def api_delete_site(name: str):
    db.delete_site(name)
    return {"ok": True}


@app.post("/api/probe")
def api_probe(payload: dict = Body(...)):
    try:
        cfg = schema.from_dict(payload)
    except Exception as e:
        return {"ok": False, "error": str(e)}
    try:
        return {"ok": True, **runner.probe_config(cfg)}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/run")
def api_run(payload: dict = Body(...)):
    name = (payload or {}).get("name")
    cfg_dict = db.get_site(name)
    if not cfg_dict:
        return {"ok": False, "error": "save the site before running"}
    if not _start_bg(schema.from_dict(cfg_dict)):
        return {"ok": False, "error": "already running"}
    return {"ok": True}


@app.post("/api/run_all")
def api_run_all():
    started = []
    for name in db.enabled_sites():
        if _start_bg(schema.from_dict(db.get_site(name))):
            started.append(name)
    return {"ok": True, "started": started}


@app.post("/api/stop")
def api_stop(payload: dict = Body(...)):
    name = (payload or {}).get("name")
    with _lock:
        live = name in _running
    if live:
        with _lock:
            _cancel.add(name)
        return {"ok": True, "mode": "graceful"}
    # no live thread for it -> stale/zombie row; force-finish it in the DB
    n = db.force_stop_site_runs(name)
    if n:
        return {"ok": True, "mode": "forced", "stopped": n}
    return {"ok": False, "error": "not running"}


# ---- export --------------------------------------------------------------
def _export(site: str, dedup: str = ""):
    members, grp, prefix = db._resolve(site)
    if grp:
        fields = (["_category"] + db.columns_for(members[0])) if members else []
        rows = []
        for cat, data in db._merged_records(members, prefix, dedup or None):
            d = dict(data); d["_category"] = cat; rows.append(d)
        return fields, rows
    fields = db.site_fields(site) or []
    rows = [r["data"] for r in db.export_rows(site)]
    if not fields and rows:
        fields = sorted({k for r in rows for k in r if not k.startswith("_")})
    return fields, rows


@app.get("/export.csv")
def export_csv(site: str, dedup: str = ""):
    fields, rows = _export(site, dedup)
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(fields)
    for r in rows:
        w.writerow(["" if r.get(k) is None else r.get(k) for k in fields])
    buf.seek(0)
    label = (site[6:] + "_merged") if site.startswith("group:") else site
    fn = f"{label}_{datetime.date.today():%Y%m%d}.csv"
    return StreamingResponse(iter([buf.getvalue()]), media_type="text/csv",
                             headers={"Content-Disposition": f"attachment; filename={fn}"})


@app.get("/export.xlsx")
def export_xlsx(site: str):
    fields, rows = _export(site)
    wb = Workbook()
    ws = wb.active
    ws.title = (site or "data")[:31]
    ws.append(fields)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for r in rows:
        ws.append(["" if r.get(k) is None else str(r.get(k)) for k in fields])
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = f"{site}_{datetime.date.today():%Y%m%d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fn}"})


@app.get("/export_all.xlsx")
def export_all_xlsx():
    """One workbook, one sheet per site — every site's data in a single file."""
    wb = Workbook()
    wb.remove(wb.active)
    used_titles = set()
    for s in db.sites_overview():
        name = s["name"]
        fields, rows = _export(name)
        # sheet titles: <=31 chars, none of []:*?/\, and unique
        title = name[:31] or "site"
        for bad in '[]:*?/\\':
            title = title.replace(bad, "_")
        base, i = title, 2
        while title in used_titles:
            title = f"{base[:28]}_{i}"
            i += 1
        used_titles.add(title)
        ws = wb.create_sheet(title=title)
        ws.append(fields or ["(no records yet)"])
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for r in rows:
            ws.append(["" if r.get(k) is None else str(r.get(k)) for k in fields])
        ws.freeze_panes = "A2"
    if not wb.sheetnames:
        wb.create_sheet(title="empty")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = f"all_sites_{datetime.date.today():%Y%m%d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fn}"})


@app.get("/api/grid")
def api_grid(site: str, offset: int = 0, limit: int = 100, sort: str = "",
             dir: str = "asc", fcol: str = "", fop: str = "", fval: str = ""):
    return db.grid(site, offset, limit, sort, dir, fcol, fop, fval)


@app.get("/api/analyze")
def api_analyze(site: str, col: str, top: int = 50, search: str = ""):
    return db.analyze(site, col, top, search)


@app.post("/api/recategorize")
def api_recategorize(payload: dict = Body(...)):
    p = payload or {}
    n = db.recategorize(p.get("site"), p.get("col"), p.get("old_value", ""),
                        p.get("new_value", ""), bool(p.get("is_blank")))
    return {"ok": True, "updated": n}


@app.post("/api/delete_category")
def api_delete_category(payload: dict = Body(...)):
    p = payload or {}
    n = db.delete_category(p.get("site"), p.get("col"), p.get("value", ""),
                           bool(p.get("is_blank")))
    return {"ok": True, "deleted": n}


@app.get("/export.db")
def export_db(site: str, dedup: str = ""):
    fd, path = tempfile.mkstemp(suffix=".db")
    os.close(fd)
    db.export_db_file(site, path, dedup)
    label = (site[6:] + "_merged") if site.startswith("group:") else site
    fn = f"{label}_{datetime.date.today():%Y%m%d}.db"
    return FileResponse(path, filename=fn, media_type="application/x-sqlite3",
                        background=BackgroundTask(os.remove, path))


@app.get("/export_all.db")
def export_all_db():
    fd, path = tempfile.mkstemp(suffix=".db")
    os.close(fd)
    db.export_all_db_file(path)
    fn = f"all_sites_{datetime.date.today():%Y%m%d}.db"
    return FileResponse(path, filename=fn, media_type="application/x-sqlite3",
                        background=BackgroundTask(os.remove, path))


@app.get("/data", response_class=HTMLResponse)
def data_page():
    return DATA_PAGE


@app.get("/", response_class=HTMLResponse)
def index():
    return PAGE


PAGE = r"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Scraper monitor</title>
<style>
  :root{
    --bg:#0e1116; --panel:#161b22; --panel-2:#1c232d; --line:#2a3340;
    --ink:#e6edf3; --muted:#8b97a6; --dim:#5c6776;
    --live:#3fb950; --live-dim:#1f3d27; --warn:#d29922; --bad:#f85149; --bar:#2f81f7;
    --mono:"SF Mono",ui-monospace,"JetBrains Mono",Menlo,Consolas,monospace;
  }
  *{box-sizing:border-box}
  body{margin:0;background:var(--bg);color:var(--ink);
    font-family:var(--mono);font-size:13px;line-height:1.5;-webkit-font-smoothing:antialiased}
  a{color:var(--bar);text-decoration:none}
  .wrap{max-width:1180px;margin:0 auto;padding:28px 22px 80px}
  header{display:flex;align-items:baseline;gap:14px;flex-wrap:wrap;
    border-bottom:1px solid var(--line);padding-bottom:16px;margin-bottom:22px}
  h1{font-size:15px;font-weight:600;letter-spacing:.04em;margin:0;text-transform:uppercase}
  .sub{color:var(--dim);font-size:12px}
  .live{margin-left:auto;display:flex;align-items:center;gap:8px;color:var(--muted);font-size:12px}
  .dot{width:8px;height:8px;border-radius:50%;background:var(--live);
    box-shadow:0 0 0 0 var(--live);animation:pulse 1.8s infinite}
  .dot.off{background:var(--dim);animation:none;box-shadow:none}
  @keyframes pulse{0%{box-shadow:0 0 0 0 rgba(63,185,80,.5)}
    70%{box-shadow:0 0 0 7px rgba(63,185,80,0)}100%{box-shadow:0 0 0 0 rgba(63,185,80,0)}}
  .grid{display:grid;grid-template-columns:repeat(12,1fr);gap:14px}
  .card{background:var(--panel);border:1px solid var(--line);border-radius:8px;padding:16px 18px}
  .card h2{font-size:11px;letter-spacing:.08em;text-transform:uppercase;color:var(--muted);
    margin:0 0 12px;font-weight:600;display:flex;align-items:center;gap:10px}
  .span6{grid-column:span 6}.span12{grid-column:span 12}
  @media(max-width:780px){.span6{grid-column:span 12}}
  .sites{display:grid;grid-template-columns:1fr 1fr;gap:10px;max-height:560px;overflow-y:auto;padding-right:4px}
  #runs{max-height:560px;overflow-y:auto;padding-right:4px}
  @media(max-width:560px){.sites{grid-template-columns:1fr}}
  .site{border:1px solid var(--line);border-radius:7px;padding:11px 13px;cursor:pointer;transition:border-color .15s,background .15s}
  .site:hover{border-color:var(--bar);background:var(--panel-2)}
  .site.sel{border-color:var(--bar);background:#202733}
  .site .nm{font-weight:600;color:var(--ink)}
  .site .meta{display:flex;justify-content:space-between;margin-top:6px;color:var(--muted)}
  .site .big{font-size:22px;font-weight:600;font-variant-numeric:tabular-nums}
  .pill{display:inline-block;padding:1px 7px;border-radius:20px;font-size:11px;border:1px solid var(--line);text-transform:lowercase}
  .pill.ok{color:var(--live);border-color:var(--live-dim);background:var(--live-dim)}
  .pill.running{color:var(--bar);border-color:var(--bar)}
  .pill.error{color:var(--bad);border-color:var(--bad)}
  .pill.partial{color:var(--warn);border-color:var(--warn)}
  .pill.interrupted{color:var(--dim);border-color:var(--line)}
  table{width:100%;border-collapse:collapse;font-size:12px}
  th{text-align:left;color:var(--muted);font-weight:600;font-size:11px;letter-spacing:.04em;
    text-transform:uppercase;padding:8px 10px;border-bottom:1px solid var(--line);position:sticky;top:0;background:var(--panel)}
  td{padding:7px 10px;border-bottom:1px solid var(--panel-2);vertical-align:top;max-width:280px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
  .tablewrap{max-height:460px;overflow:auto;border:1px solid var(--line);border-radius:8px}
  .fills{display:flex;flex-direction:column;gap:9px}
  .fill .top{display:flex;justify-content:space-between;align-items:baseline}
  .fill .nm{color:var(--ink)} .fill .pc{font-variant-numeric:tabular-nums;color:var(--muted)}
  .fbar{height:6px;background:var(--panel-2);border-radius:4px;overflow:hidden;margin-top:4px;border:1px solid var(--line)}
  .fbar>span{display:block;height:100%;background:var(--live)}
  .fill.drift .nm,.fill.drift .pc{color:var(--bad)} .fill.drift .fbar>span{background:var(--bad)}
  .driftnote{color:var(--bad);font-size:11px;margin-top:2px}
  .runrow{display:flex;align-items:center;gap:10px;padding:7px 0;border-top:1px solid var(--line)}
  .runrow:first-child{border-top:none}
  .runrow .s{flex:1;color:var(--muted)} .runrow .n{color:var(--ink);font-variant-numeric:tabular-nums}
  .runrow .t{color:var(--dim);font-size:11px;min-width:84px;text-align:right}
  .controls{display:flex;gap:8px;align-items:center;margin-bottom:12px;flex-wrap:wrap}
  .actions{display:flex;gap:10px;flex-wrap:wrap;align-items:center}
  .btn{display:inline-flex;align-items:center;gap:7px;background:var(--panel-2);border:1px solid var(--line);
    color:var(--ink);padding:8px 13px;border-radius:7px;font-family:var(--mono);font-size:12px;cursor:pointer}
  .btn:hover{border-color:var(--bar);background:#202733}
  .btn.primary{border-color:var(--bar);color:#fff;background:#1b3a63}
  .btn.danger:hover{border-color:var(--bad);color:var(--bad)}
  .btn:disabled{opacity:.5;cursor:not-allowed}
  label{display:block;color:var(--muted);font-size:11px;text-transform:uppercase;letter-spacing:.04em;margin:0 0 4px}
  input,select,textarea{background:var(--panel-2);border:1px solid var(--line);color:var(--ink);
    padding:7px 9px;border-radius:6px;font-family:var(--mono);font-size:12px;width:100%}
  textarea{resize:vertical;min-height:84px;white-space:pre}
  .formgrid{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-bottom:14px}
  @media(max-width:780px){.formgrid{grid-template-columns:1fr 1fr}}
  .ftable{width:100%;border-collapse:collapse;margin-bottom:10px}
  .ftable th{padding:5px 6px;font-size:10px}
  .ftable td{padding:4px 6px;border:none;max-width:none;white-space:normal;overflow:visible}
  .ftable input,.ftable select{padding:5px 7px}
  .xbtn{background:none;border:1px solid var(--line);color:var(--dim);border-radius:5px;cursor:pointer;padding:4px 8px}
  .xbtn:hover{color:var(--bad);border-color:var(--bad)}
  .result{margin-top:12px;font-size:12px;border-top:1px solid var(--line);padding-top:10px}
  .res-hit{color:var(--live)} .res-miss{color:var(--bad)}
  .res-line{display:flex;gap:8px;padding:2px 0}
  .res-line .k{min-width:120px;color:var(--muted)}
  .msg{font-size:12px;margin-left:auto}
  .msg.ok{color:var(--live)} .msg.err{color:var(--bad)}
  input{background:var(--panel-2)}
  .empty{color:var(--dim);padding:8px 2px}
  .jsononly{display:none}
</style>
</head>
<body>
<div class="wrap">
  <header>
    <h1>Scraper monitor</h1>
    <span class="sub">config-driven</span>
    <a class="btn" href="/data" style="padding:4px 10px">data grid ↗</a>
    <span class="live"><span id="dot" class="dot"></span><span id="livetxt">live</span></span>
  </header>

  <div class="grid">
    <div class="card span6">
      <h2>Sites <button class="btn" id="newbtn" style="margin-left:auto;padding:4px 10px">+ new</button>
        <button class="btn" id="runallbtn" style="padding:4px 10px">run all</button>
        <a class="btn" href="/export_all.xlsx" style="padding:4px 10px" title="every site, one sheet each">⬇ all xlsx</a></h2>
      <div class="sites" id="sites"></div>
    </div>

    <div class="card span6">
      <h2>Recent runs <button class="btn" id="clrruns" style="margin-left:auto;padding:4px 10px">clear</button></h2>
      <div id="runs"></div>
    </div>

    <div class="card span12">
      <h2>Configuration <span class="sub">— edit everything here, no files</span>
        <span class="msg" id="edmsg"></span></h2>
      <div class="formgrid">
        <div><label>name (unique)</label><input id="f-name" placeholder="my_site"></div>
        <div><label>engine</label>
          <select id="f-engine">
            <option value="auto">auto</option>
            <option value="http_html">http_html</option>
            <option value="browser">browser</option>
            <option value="http_json">http_json</option>
          </select></div>
        <div style="grid-column:span 2"><label>url</label><input id="f-url" placeholder="https://…"></div>
        <div><label>key fields (dedup, comma)</label><input id="f-key" placeholder="url"></div>
        <div class="htmlonly"><label>list.container</label><input id="f-container" placeholder="article.product"></div>
        <div class="htmlonly"><label>next_page selector</label><input id="f-next" placeholder="li.next a::attr(href)"></div>
        <div class="htmlonly"><label>wait_for (browser)</label><input id="f-wait" placeholder=".loaded"></div>
        <div class="jsononly"><label>records_path</label><input id="f-recpath" placeholder="data"></div>
        <div class="jsononly"><label>total_path</label><input id="f-totpath" placeholder="totalCount"></div>
      </div>

      <table class="ftable">
        <thead><tr><th style="width:18%">field</th><th style="width:30%">selector / path</th>
          <th style="width:12%">type</th><th style="width:12%">attr</th>
          <th style="width:12%">transform</th><th style="width:12%">match/regex</th><th></th></tr></thead>
        <tbody id="frows"></tbody>
      </table>
      <button class="btn" id="addfield" style="padding:5px 11px">+ field</button>

      <div class="jsononly" style="margin-top:14px">
        <label>request block (JSON — method/url/body/page params)</label>
        <textarea id="f-request" placeholder='{"method":"POST","url":"…","body":{},"page_param":"page","page_size_param":"pageSize","page_size":25}'></textarea>
      </div>

      <div class="actions" style="margin-top:16px">
        <button class="btn primary" id="savebtn">save</button>
        <button class="btn" id="probebtn">probe</button>
        <button class="btn" id="runbtn">save &amp; run</button>
        <button class="btn danger" id="delbtn" style="margin-left:auto">delete</button>
      </div>
      <div class="result" id="probe-result" style="display:none"></div>
    </div>

    <div class="card span6">
      <h2>Field fill rate <span class="sub">— <span id="fr-site">select a site</span></span></h2>
      <div class="fills" id="fills"><div class="empty">pick a site</div></div>
    </div>

    <div class="card span6">
      <h2>Export <span class="sub">— <span id="ex-site"></span></span></h2>
      <p class="sub" style="margin:0 0 14px">Everything stored for this site. Always current.</p>
      <div class="actions" id="exports"><div class="empty">pick a site</div></div>
    </div>

    <div class="card span12">
      <h2>Recent records <span class="sub">— <span id="pv-site"></span></span></h2>
      <div class="controls">
        <input id="q" placeholder="filter records…" style="width:auto">
        <span class="sub">showing latest <span id="pvcount">0</span></span>
        <button class="btn" id="delsel" style="padding:5px 11px;margin-left:auto">delete selected</button>
        <button class="btn danger" id="clrsite" style="padding:5px 11px">clear site</button>
        <button class="btn danger" id="clrall" style="padding:5px 11px">clear all</button>
      </div>
      <div class="tablewrap"><table><thead id="thead"></thead><tbody id="tbody"></tbody></table></div>
    </div>
  </div>
  <div class="msg err" id="err" style="margin-top:10px"></div>
</div>

<script>
const $ = id => document.getElementById(id);
let SITE = null, alive = true;
const TYPES = ["string","number","boolean","url","list"];
const TRANSFORMS = ["","currency","int","lower","upper","strip"];
const fmtPct = r => Math.round((r||0)*100)+'%';
const ago = ts => { if(!ts) return '—'; const s=(Date.now()-new Date(ts).getTime())/1000;
  if(s<60) return Math.round(s)+'s ago'; if(s<3600) return Math.round(s/60)+'m ago';
  if(s<86400) return Math.round(s/3600)+'h ago'; return Math.round(s/86400)+'d ago'; };
const flash = (el,msg,ok)=>{ el.textContent=msg; el.className='msg '+(ok?'ok':'err');
  setTimeout(()=>{ if(el.textContent===msg) el.textContent=''; }, 4000); };

/* ---------- editor ---------- */
function toggleEngineBlocks(){
  const json = $('f-engine').value==='http_json';
  document.querySelectorAll('.jsononly').forEach(e=>e.style.display=json?'':'none');
  document.querySelectorAll('.htmlonly').forEach(e=>e.style.display=json?'none':'');
}
function fieldRow(name='', spec={}){
  const tr=document.createElement('tr');
  const sel = spec.selector!==undefined?spec.selector:(spec.path||'');
  const mr = spec.match!==undefined?spec.match:(spec.regex||'');
  tr.innerHTML=`
    <td><input class="r-name" value="${name}"></td>
    <td><input class="r-sel" value="${(sel||'').replace(/"/g,'&quot;')}"></td>
    <td><select class="r-type">${TYPES.map(t=>`<option ${t===(spec.type||'string')?'selected':''}>${t}</option>`).join('')}</select></td>
    <td><input class="r-attr" value="${spec.attr||''}" placeholder="href"></td>
    <td><select class="r-tf">${TRANSFORMS.map(t=>`<option ${t===(spec.transform||'')?'selected':''}>${t||'—'}</option>`).join('')}</select></td>
    <td><input class="r-mr" value="${(mr||'').replace(/"/g,'&quot;')}" placeholder="match or regex"></td>
    <td><button class="xbtn">×</button></td>`;
  tr.querySelector('.xbtn').onclick=()=>tr.remove();
  return tr;
}
function clearEditor(){
  ['f-name','f-url','f-key','f-container','f-next','f-wait','f-recpath','f-totpath','f-request'].forEach(i=>$(i).value='');
  $('f-engine').value='auto'; toggleEngineBlocks();
  $('frows').innerHTML=''; $('frows').appendChild(fieldRow());
  $('probe-result').style.display='none';
}
function loadEditor(cfg){
  if(!cfg){ clearEditor(); return; }
  $('f-name').value=cfg.name||''; $('f-url').value=cfg.url||'';
  $('f-engine').value=cfg.engine||'auto'; toggleEngineBlocks();
  $('f-key').value=(cfg.key_fields||[]).join(', ');
  $('f-container').value=(cfg.list&&cfg.list.container)||'';
  $('f-next').value=cfg.next_page||''; $('f-wait').value=cfg.wait_for||'';
  $('f-recpath').value=cfg.records_path||''; $('f-totpath').value=cfg.total_path||'';
  $('f-request').value=cfg.request?JSON.stringify(cfg.request,null,2):'';
  const tb=$('frows'); tb.innerHTML='';
  Object.entries(cfg.fields||{}).forEach(([n,s])=>tb.appendChild(fieldRow(n,s)));
  if(!Object.keys(cfg.fields||{}).length) tb.appendChild(fieldRow());
  $('probe-result').style.display='none';
}
function buildConfig(){
  const engine=$('f-engine').value, json=engine==='http_json';
  const cfg={name:$('f-name').value.trim(), url:$('f-url').value.trim(), engine};
  const key=$('f-key').value.trim(); if(key) cfg.key_fields=key.split(',').map(s=>s.trim()).filter(Boolean);
  if(json){
    cfg.records_path=$('f-recpath').value.trim();
    const tot=$('f-totpath').value.trim(); if(tot) cfg.total_path=tot;
    const reqtxt=$('f-request').value.trim();
    if(reqtxt){ cfg.request=JSON.parse(reqtxt); }   // throws -> caught by caller
  } else {
    const c=$('f-container').value.trim(); if(c) cfg.list={container:c};
    const n=$('f-next').value.trim(); if(n) cfg.next_page=n;
    const w=$('f-wait').value.trim(); if(w) cfg.wait_for=w;
  }
  cfg.fields={};
  $('frows').querySelectorAll('tr').forEach(tr=>{
    const nm=tr.querySelector('.r-name').value.trim(); if(!nm) return;
    const spec={type:tr.querySelector('.r-type').value};
    const sel=tr.querySelector('.r-sel').value.trim();
    if(json) spec.path=sel; else spec.selector=sel;
    const attr=tr.querySelector('.r-attr').value.trim(); if(attr) spec.attr=attr;
    const tf=tr.querySelector('.r-tf').value.trim(); if(tf&&tf!=='—') spec.transform=tf;
    const mr=tr.querySelector('.r-mr').value.trim();
    if(mr){ if(spec.type==='boolean') spec.match=mr; else spec.regex=mr; }
    cfg.fields[nm]=spec;
  });
  return cfg;
}

async function save(){
  let cfg; try{ cfg=buildConfig(); }catch(e){ flash($('edmsg'),'bad request JSON',false); return null; }
  const r=await (await fetch('/api/site',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(cfg)})).json();
  if(r.ok){ flash($('edmsg'),'saved',true); SITE=r.name; loadSites(); }
  else flash($('edmsg'), r.error||'save failed', false);
  return r.ok?cfg:null;
}
async function probe(){
  let cfg; try{ cfg=buildConfig(); }catch(e){ flash($('edmsg'),'bad request JSON',false); return; }
  const box=$('probe-result'); box.style.display='block';
  box.innerHTML='<span class="sub">probing… one fetch, this can take a few seconds</span>';
  const r=await (await fetch('/api/probe',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(cfg)})).json();
  if(!r.ok){ box.innerHTML=`<span class="res-miss">error: ${r.error}</span>`; return; }
  const head=`<div style="margin-bottom:8px">engine <b>${r.engine_used}</b> · ${r.total} record(s) on first page</div>`;
  const lines=r.fields.map(f=>{
    let v=f.value; if(Array.isArray(v))v=`[${v.length}] `+JSON.stringify(v.slice(0,2));
    else if(v===null||v===undefined)v='—'; else v=String(v);
    return `<div class="res-line"><span class="k ${f.hit?'res-hit':'res-miss'}">${f.hit?'HIT':'MISS'} ${f.field}</span><span>${v.slice(0,160).replace(/</g,'&lt;')}</span></div>`;
  }).join('');
  const miss=r.fields.filter(f=>!f.hit).map(f=>f.field);
  box.innerHTML=head+lines+(miss.length?`<div class="driftnote">fix these selectors: ${miss.join(', ')}</div>`:'<div class="res-hit" style="margin-top:6px">all fields hit — safe to run</div>');
}
async function runNow(){
  const cfg=await save(); if(!cfg) return;
  const r=await (await fetch('/api/run',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({name:cfg.name})})).json();
  flash($('edmsg'), r.ok?'run started — watch it live':(r.error||'could not start'), r.ok);
  setTimeout(()=>{loadSites();loadRuns();}, 600);
}
async function del(){
  const name=$('f-name').value.trim(); if(!name) return;
  if(!confirm('Delete '+name+' and all its records?')) return;
  await fetch('/api/site?name='+encodeURIComponent(name),{method:'DELETE'});
  if(SITE===name) SITE=null; clearEditor(); loadSites(); flash($('edmsg'),'deleted',true);
}

$('newbtn').onclick=()=>{ SITE=null; clearEditor(); loadSites(); };
$('addfield').onclick=()=>$('frows').appendChild(fieldRow());
$('f-engine').onchange=toggleEngineBlocks;
$('savebtn').onclick=save; $('probebtn').onclick=probe;
$('runbtn').onclick=runNow; $('delbtn').onclick=del;
$('runallbtn').onclick=async()=>{ await fetch('/api/run_all',{method:'POST'}); setTimeout(()=>{loadSites();loadRuns();},600); };

/* ---------- live reads ---------- */
async function loadSites(){
  try{
    const s=await (await fetch('/api/sites')).json();
    $('err').textContent='';
    const box=$('sites'); box.innerHTML='';
    if(!s.sites.length){ box.innerHTML='<div class="empty">no sites — click <b>+ new</b> to add one</div>'; }
    const running=new Set(s.running||[]);
    s.sites.forEach(x=>{
      const el=document.createElement('div'); el.className='site'+(x.name===SITE?' sel':'');
      const st=running.has(x.name)?'running':(x.last_status||'—');
      el.innerHTML=`<div class="nm">${x.name}</div><div class="big">${(x.records||0).toLocaleString()}</div>
        <div class="meta"><span class="pill ${st}">${st}</span><span>${ago(x.last_run)}</span></div>`;
      el.onclick=()=>{ SITE=x.name; selectSite(); fetch('/api/site?name='+encodeURIComponent(x.name))
        .then(r=>r.json()).then(d=>loadEditor(d.config)); loadSites(); };
      box.appendChild(el);
    });
    $('dot').classList.toggle('off', !running.size && document.hidden);
  }catch(e){ $('err').textContent='backend unreachable — is Postgres up?'; }
}
function selectSite(){
  ['fr-site','ex-site','pv-site'].forEach(id=>$(id).textContent=SITE||'');
  if(SITE) $('exports').innerHTML=
    `<a class="btn" href="/export.xlsx?site=${encodeURIComponent(SITE)}">xlsx</a>
     <a class="btn" href="/export.csv?site=${encodeURIComponent(SITE)}">csv</a>`;
  loadStats(); loadPreview();
}
async function loadRuns(){
  try{
    const s=await (await fetch('/api/runs?limit=18')).json();
    const box=$('runs'); box.innerHTML=''; let running=false;
    if(!s.runs.length) box.innerHTML='<div class="empty">no runs yet</div>';
    s.runs.forEach(r=>{ if(r.status==='running')running=true;
      const el=document.createElement('div'); el.className='runrow';
      const stop = r.status==='running'
        ? `<button class="xbtn" data-stop="${r.site}" style="margin-left:6px">stop</button>` : '';
      el.innerHTML=`<span class="pill ${r.status}">${r.status}</span><span class="s">${r.site}</span>
        <span class="n">${(r.records_found||0).toLocaleString()} found · ${(r.records_new||0).toLocaleString()} new</span>
        <span class="t">${r.secs?Math.round(r.secs)+'s':''} · ${ago(r.started_at)}</span>${stop}`;
      box.appendChild(el); });
    box.querySelectorAll('[data-stop]').forEach(b=>b.onclick=()=>stopRun(b.dataset.stop));
  }catch(e){}
}
async function stopRun(name){
  await fetch('/api/stop',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({name})});
  flash($('edmsg'), 'stopping '+name+' after the current page…', true);
  setTimeout(loadRuns, 600);
}
async function loadStats(){
  if(!SITE) return;
  try{
    const s=await (await fetch('/api/stats?site='+encodeURIComponent(SITE))).json();
    const box=$('fills'); box.innerHTML='';
    if(!s.fields.length){ box.innerHTML='<div class="empty">no completed run yet</div>'; return; }
    s.fields.forEach(f=>{ const el=document.createElement('div'); el.className='fill'+(f.drift?' drift':'');
      el.innerHTML=`<div class="top"><span class="nm">${f.field}</span>
        <span class="pc">${fmtPct(f.fill_rate)} · ${f.filled}/${f.total}</span></div>
        <div class="fbar"><span style="width:${Math.round((f.fill_rate||0)*100)}%"></span></div>
        ${f.drift?`<div class="driftnote">drift: was ~${fmtPct(f.avg_rate)}, now ${fmtPct(f.fill_rate)} — selector likely broke</div>`:''}`;
      box.appendChild(el); });
  }catch(e){}
}
async function loadPreview(){
  if(!SITE) return;
  try{
    const q=$('q').value.trim(); const u=new URLSearchParams({site:SITE,limit:60}); if(q)u.set('q',q);
    const s=await (await fetch('/api/preview?'+u)).json();
    const cols=s.fields||[];
    $('thead').innerHTML='<tr><th style="width:26px"><input type="checkbox" id="selall"></th>'+cols.map(c=>`<th>${c}</th>`).join('')+'</tr>';
    const tb=$('tbody'); tb.innerHTML='';
    (s.rows||[]).forEach(row=>{
      const tr=document.createElement('tr');
      const cb=`<td><input type="checkbox" class="rsel" value="${row._rid}"></td>`;
      tr.innerHTML=cb+cols.map(c=>{ let v=row[c];
        if(v===null||v===undefined)v='<span style="color:var(--dim)">—</span>';
        else if(Array.isArray(v))v=v.join(', ');
        else if(typeof v==='string'&&/^https?:\/\//.test(v))v=`<a href="${v}" target="_blank" rel="noopener">link</a>`;
        else if(v===true)v='<span style="color:var(--live)">yes</span>';
        else if(v===false)v='<span style="color:var(--dim)">no</span>';
        return `<td title="${String(row[c]??'').replace(/"/g,'&quot;')}">${v}</td>`; }).join('');
      tb.appendChild(tr); });
    $('pvcount').textContent=(s.rows||[]).length;
    const sa=$('selall'); if(sa) sa.onclick=()=>document.querySelectorAll('.rsel').forEach(c=>c.checked=sa.checked);
  }catch(e){}
}

function tick(){ loadSites(); loadRuns(); loadStats(); loadPreview(); }
$('q').oninput=(()=>{let t;return()=>{clearTimeout(t);t=setTimeout(loadPreview,300);};})();

async function delSelected(){
  const ids=[...document.querySelectorAll('.rsel:checked')].map(c=>c.value);
  if(!ids.length){ flash($('edmsg'),'no rows selected',false); return; }
  if(!confirm(`Delete ${ids.length} selected record(s)? The config stays.`)) return;
  const r=await (await fetch(`/api/records?site=${encodeURIComponent(SITE)}&ids=${ids.join(',')}`,{method:'DELETE'})).json();
  flash($('edmsg'),`deleted ${r.deleted} record(s)`,true); loadPreview(); loadSites();
}
async function clearSite(){
  if(!SITE) return;
  if(!confirm(`Delete ALL records for ${SITE}? The config stays, the data is wiped.`)) return;
  const r=await (await fetch(`/api/records?site=${encodeURIComponent(SITE)}`,{method:'DELETE'})).json();
  flash($('edmsg'),`cleared ${r.deleted} record(s) from ${SITE}`,true); loadPreview(); loadSites();
}
async function clearAll(){
  if(!confirm('Delete ALL scraped data for EVERY site? Configs stay; all records go.')) return;
  if(!confirm('Final check — this wipes every record across all sites. Continue?')) return;
  const r=await (await fetch('/api/records?all=true',{method:'DELETE'})).json();
  flash($('edmsg'),`cleared ${r.deleted} record(s) across all sites`,true); loadPreview(); loadSites();
}
$('delsel').onclick=delSelected; $('clrsite').onclick=clearSite; $('clrall').onclick=clearAll;
async function clearRuns(){
  if(!confirm('Clear finished run history? Any run still in progress is kept.')) return;
  const r=await (await fetch('/api/runs',{method:'DELETE'})).json();
  flash($('edmsg'),`cleared ${r.deleted} run(s)`,true); loadRuns(); loadStats();
}
$('clrruns').onclick=clearRuns;
clearEditor(); tick();
setInterval(()=>{ if(alive){ loadSites(); loadRuns(); loadStats(); loadPreview(); } }, 4000);
document.addEventListener('visibilitychange',()=>{ alive=!document.hidden;
  $('livetxt').textContent=document.hidden?'paused':'live'; if(!document.hidden) tick(); });
</script>
</body>
</html>"""


DATA_PAGE = r"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Data grid</title>
<style>
  :root{
    --bg:#0e1116; --panel:#161b22; --panel-2:#1c232d; --line:#2a3340;
    --ink:#e6edf3; --muted:#8b97a6; --dim:#5c6776; --bar:#2f81f7; --live:#3fb950; --bad:#f85149;
    --mono:"SF Mono",ui-monospace,"JetBrains Mono",Menlo,Consolas,monospace;
  }
  *{box-sizing:border-box}
  body{margin:0;background:var(--bg);color:var(--ink);font-family:var(--mono);font-size:13px}
  a{color:var(--bar);text-decoration:none}
  .wrap{max-width:1500px;margin:0 auto;padding:20px 18px 60px}
  header{display:flex;align-items:baseline;gap:14px;border-bottom:1px solid var(--line);padding-bottom:14px;margin-bottom:16px}
  h1{font-size:15px;font-weight:600;letter-spacing:.04em;margin:0;text-transform:uppercase}
  .controls{display:flex;gap:8px;flex-wrap:wrap;align-items:center;margin-bottom:12px}
  select,input{background:var(--panel-2);border:1px solid var(--line);color:var(--ink);
    padding:7px 9px;border-radius:6px;font-family:var(--mono);font-size:12px}
  .btn{display:inline-flex;align-items:center;gap:6px;background:var(--panel-2);border:1px solid var(--line);
    color:var(--ink);padding:7px 12px;border-radius:6px;font-family:var(--mono);font-size:12px;cursor:pointer}
  .btn:hover{border-color:var(--bar);background:#202733}
  .btn.danger:hover{border-color:var(--bad);color:var(--bad)}
  .spacer{flex:1}
  .tablewrap{border:1px solid var(--line);border-radius:8px;overflow:auto;max-height:72vh}
  table{width:100%;border-collapse:collapse;font-size:12px}
  th{position:sticky;top:0;background:var(--panel);text-align:left;color:var(--muted);font-weight:600;
    font-size:11px;letter-spacing:.03em;text-transform:uppercase;padding:8px 10px;border-bottom:1px solid var(--line);white-space:nowrap;cursor:pointer;user-select:none}
  th.nosort{cursor:default}
  td{padding:6px 10px;border-bottom:1px solid var(--panel-2);max-width:340px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
  tr:hover td{background:var(--panel-2)}
  .pager{display:flex;gap:10px;align-items:center;margin-top:12px;color:var(--muted);flex-wrap:wrap}
  .muted{color:var(--dim)} .live{color:var(--live)}
  .toast{position:fixed;bottom:18px;left:50%;transform:translateX(-50%);background:var(--panel);
    border:1px solid var(--line);padding:9px 16px;border-radius:7px;opacity:0;transition:opacity .2s;font-size:12px}
  .toast.show{opacity:1}
  .tabs{display:flex;gap:8px;margin:4px 0 10px}
  .tab-active{border-color:var(--bar);background:#202733}
  .chartwrap{background:var(--panel);border:1px solid var(--line);border-radius:8px;padding:14px 16px;margin-top:8px}
  .chartwrap h3{font-size:12px;color:var(--muted);font-weight:600;margin:0 0 12px}
  .barrow{display:flex;align-items:center;gap:10px;margin:5px 0}
  .barlabel{width:200px;flex-shrink:0;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
  .bartrack{flex:1;height:14px;background:var(--panel-2);border:1px solid var(--line);border-radius:4px;overflow:hidden}
  .barfill{height:100%;background:var(--bar)}
  .barval{width:250px;flex-shrink:0;text-align:right;color:var(--muted)}
  .mini{background:var(--panel-2);border:1px solid var(--line);color:var(--ink);border-radius:5px;padding:2px 7px;font-family:var(--mono);font-size:11px;cursor:pointer;margin-left:6px}
  .mini:hover{border-color:var(--bar)} .mini.danger:hover{border-color:var(--bad);color:var(--bad)}
</style>
</head>
<body>
<div class="wrap">
  <header>
    <h1>Data grid</h1><a href="/">&larr; back to monitor</a>
    <span class="spacer"></span>
    <span class="muted" id="rowinfo"></span>
  </header>

  <div class="tabs">
    <button class="btn tab-active" id="tabgrid">grid</button>
    <button class="btn" id="tabanalyze">analyze</button>
  </div>

  <div id="gridview">
  <div class="controls">
    <label class="muted">site</label>
    <select id="site"></select>
    <span class="spacer"></span>
    <select id="fcol"></select>
    <select id="fop">
      <option value="contains">contains</option>
      <option value="equals">equals</option>
      <option value="starts_with">starts with</option>
      <option value="not_equals">not equals</option>
      <option value="greater">&gt;</option>
      <option value="less">&lt;</option>
    </select>
    <input id="fval" placeholder="filter value…" size="18">
    <button class="btn" id="apply">filter</button>
    <button class="btn" id="reset">reset</button>
  </div>
  <div class="controls">
    <button class="btn danger" id="delsel">delete selected</button>
    <button class="btn danger" id="dedupe">remove duplicate titles</button>
    <span class="spacer"></span>
    <label id="dedupwrap" class="muted" style="display:none"><input type="checkbox" id="dedup" checked> unique books (by ISBN)</label>
    <a class="btn" id="dldb">⬇ .db (SQLite)</a>
    <a class="btn" id="dlcsv">⬇ .csv</a>
    <a class="btn" href="/export_all.db">⬇ all sites .db</a>
  </div>

  <div class="tablewrap"><table><thead id="thead"></thead><tbody id="tbody"></tbody></table></div>

  <div class="pager">
    <button class="btn" id="first">« first</button>
    <button class="btn" id="prev">‹ prev</button>
    <span id="pageinfo">—</span>
    <button class="btn" id="next">next ›</button>
    <label class="muted">rows/page</label>
    <select id="psize"><option>50</option><option selected>100</option><option>250</option><option>500</option></select>
  </div>
  </div><!-- gridview -->

  <div id="analyzeview" style="display:none">
    <div class="controls">
      <label class="muted">column</label>
      <select id="acol"></select>
      <input id="asearch" placeholder="only values containing… (optional)" size="24">
      <button class="btn" id="arun">analyze</button>
      <span class="spacer"></span>
      <span class="muted" id="ainfo"></span>
    </div>
    <div id="analyzeArea"><div class="muted">pick a column and click analyze</div></div>
  </div>
</div>
<div class="toast" id="toast"></div>

<script>
const $=s=>document.querySelector(s);
const fmt=n=>(n||0).toLocaleString();
let st={site:null,offset:0,limit:100,sort:"",dir:"asc",fcol:"__all__",fop:"contains",fval:"",total:0,columns:[]};
function toast(m){const t=$('#toast');t.textContent=m;t.classList.add('show');setTimeout(()=>t.classList.remove('show'),2200);}

async function loadSites(){
  const s=await (await fetch('/api/sites')).json();
  const sel=$('#site');
  let opts=s.sites.map(x=>`<option value="${x.name}">${x.name}</option>`);
  const grp={};
  s.sites.forEach(x=>{const p=x.name.split('_')[0];(grp[p]=grp[p]||[]).push(x.name);});
  Object.entries(grp).filter(([p,m])=>m.length>1).forEach(([p,m])=>{
    opts.unshift(`<option value="group:${p}">⊕ ${p} — all ${m.length} merged</option>`);
  });
  sel.innerHTML=opts.join('');
  if(opts.length){ st.site=sel.options[0].value; await loadGrid(); }
  else { $('#tbody').innerHTML='<tr><td class="muted">no sites yet</td></tr>'; }
}

function buildFcol(){
  $('#fcol').innerHTML=`<option value="__all__">All columns</option>`+
    st.columns.map(c=>`<option value="${c}">${c}</option>`).join('');
  $('#fcol').value=st.fcol;
}

async function loadGrid(){
  const u=new URLSearchParams({site:st.site,offset:st.offset,limit:st.limit,
    sort:st.sort,dir:st.dir,fcol:st.fcol,fop:st.fop,fval:st.fval});
  const d=await (await fetch('/api/grid?'+u)).json();
  st.total=d.total; st.columns=d.columns||[];
  buildFcol();
  const arrow=c=>st.sort===c?(st.dir==='asc'?' ▲':' ▼'):'';
  $('#thead').innerHTML='<tr><th class="nosort" style="width:26px"><input type="checkbox" id="selall"></th>'+
    st.columns.map(c=>`<th data-c="${c}">${c}${arrow(c)}</th>`).join('')+'</tr>';
  $('#tbody').innerHTML=(d.rows||[]).map(r=>'<tr><td><input type="checkbox" class="rs" value="'+r._rid+'"></td>'+
    st.columns.map(c=>{let v=r[c]; if(v===null||v===undefined)v='<span class="muted">—</span>';
      else if(typeof v==='string'&&/^https?:\/\//.test(v))v='<a href="'+v+'" target="_blank" rel="noopener">link</a>';
      else v=String(v).replace(/</g,'&lt;');
      return '<td title="'+String(r[c]??'').replace(/"/g,'&quot;')+'">'+v+'</td>';}).join('')+'</tr>').join('');
  $('#thead').querySelectorAll('th[data-c]').forEach(th=>th.onclick=()=>{
    const c=th.dataset.c; if(st.sort===c)st.dir=st.dir==='asc'?'desc':'asc'; else{st.sort=c;st.dir='asc';}
    st.offset=0; loadGrid();
  });
  const sa=$('#selall'); if(sa)sa.onclick=()=>document.querySelectorAll('.rs').forEach(c=>c.checked=sa.checked);
  const from=st.total?st.offset+1:0, to=Math.min(st.offset+st.limit,st.total);
  $('#pageinfo').textContent=`${fmt(from)}–${fmt(to)} of ${fmt(st.total)}`;
  $('#rowinfo').textContent=`${fmt(st.total)} rows in ${st.site}`;
  const grp=st.site.startsWith('group:');
  $('#dedupwrap').style.display=grp?'':'none';
  const dlu='site='+encodeURIComponent(st.site);
  const dd=(grp&&$('#dedup').checked)?'&dedup=isbn':'';
  $('#dldb').href='/export.db?'+dlu+dd; $('#dlcsv').href='/export.csv?'+dlu+dd;
}
$('#dedup').onchange=()=>{
  const grp=st.site.startsWith('group:'), dlu='site='+encodeURIComponent(st.site);
  const dd=(grp&&$('#dedup').checked)?'&dedup=isbn':'';
  $('#dldb').href='/export.db?'+dlu+dd; $('#dlcsv').href='/export.csv?'+dlu+dd;
};

$('#site').onchange=e=>{st.site=e.target.value;st.offset=0;st.sort='';st.fval='';st.fcol='__all__';$('#fval').value='';loadGrid();};
$('#apply').onclick=()=>{st.fcol=$('#fcol').value;st.fop=$('#fop').value;st.fval=$('#fval').value.trim();st.offset=0;loadGrid();};
$('#reset').onclick=()=>{st.fcol='__all__';st.fop='contains';st.fval='';$('#fval').value='';st.sort='';st.offset=0;loadGrid();};
$('#fval').addEventListener('keydown',e=>{if(e.key==='Enter')$('#apply').click();});
$('#first').onclick=()=>{st.offset=0;loadGrid();};
$('#prev').onclick=()=>{st.offset=Math.max(0,st.offset-st.limit);loadGrid();};
$('#next').onclick=()=>{if(st.offset+st.limit<st.total){st.offset+=st.limit;loadGrid();}};
$('#psize').onchange=e=>{st.limit=+e.target.value;st.offset=0;loadGrid();};
$('#delsel').onclick=async()=>{
  const ids=[...document.querySelectorAll('.rs:checked')].map(c=>c.value);
  if(!ids.length){toast('no rows selected');return;}
  if(!confirm(`Delete ${ids.length} record(s) from ${st.site}?`))return;
  const d=await (await fetch(`/api/records?site=${encodeURIComponent(st.site)}&ids=${ids.join(',')}`,{method:'DELETE'})).json();
  toast(`deleted ${d.deleted}`); loadGrid();
};
$('#dedupe').onclick=async()=>{
  if(!st.site){toast('no site');return;}
  if(!confirm(`Remove duplicate "title" rows from ${st.site}? One copy of each title is kept; the rest are permanently deleted.`))return;
  const d=await (await fetch(`/api/dedupe?site=${encodeURIComponent(st.site)}&field=title`,{method:'DELETE'})).json();
  if(d.error){toast(d.error);return;}
  toast(`removed ${d.deleted} duplicate(s)`); loadGrid();
};
function esc(s){return String(s).replace(/[&<>"]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c]));}
function pct(a,b){return b?(a/b*100):0;}
function showTab(w){
  $('#gridview').style.display=w==='grid'?'':'none';
  $('#analyzeview').style.display=w==='analyze'?'':'none';
  $('#tabgrid').classList.toggle('tab-active',w==='grid');
  $('#tabanalyze').classList.toggle('tab-active',w==='analyze');
  if(w==='analyze') $('#acol').innerHTML=st.columns.map(c=>`<option>${c}</option>`).join('');
}
$('#tabgrid').onclick=()=>showTab('grid');
$('#tabanalyze').onclick=()=>showTab('analyze');
$('#arun').onclick=runAnalyze;
$('#asearch').addEventListener('keydown',e=>{if(e.key==='Enter')runAnalyze();});
let adata=null;
async function runAnalyze(){
  const col=$('#acol').value, search=$('#asearch').value.trim();
  if(!col)return;
  $('#analyzeArea').innerHTML='<div class="muted">analyzing… (aggregating in the database)</div>';
  const u=new URLSearchParams({site:st.site,col,search,top:50});
  const d=await (await fetch('/api/analyze?'+u)).json();
  if(d.error){$('#analyzeArea').innerHTML='<div class="muted">'+d.error+'</div>';return;}
  adata=d; renderAnalyze(d);
}
function renderAnalyze(d){
  $('#ainfo').textContent=`${fmt(d.distinct)} distinct in ${fmt(d.total)} rows`;
  const max=Math.max(1,...d.rows.map(r=>r.count),d.others);
  let bars=d.rows.map((r,i)=>{
    const lbl=r.value===''?'(blank)':r.value, w=(r.count/max*100).toFixed(1);
    return `<div class="barrow"><div class="barlabel" title="${esc(lbl)}">${esc(lbl)}</div>
      <div class="bartrack"><div class="barfill" style="width:${w}%"></div></div>
      <div class="barval">${fmt(r.count)} · ${pct(r.count,d.total).toFixed(1)}%
        <button class="mini" data-edit="${i}">edit</button>
        <button class="mini danger" data-del="${i}">del</button></div></div>`;}).join('');
  if(d.others>0){const w=(d.others/max*100).toFixed(1);
    bars+=`<div class="barrow"><div class="barlabel muted">Others (${fmt(d.hidden_distinct)} more)</div>
      <div class="bartrack"><div class="barfill" style="width:${w}%;background:#3a4254"></div></div>
      <div class="barval">${fmt(d.others)} · ${pct(d.others,d.total).toFixed(1)}%</div></div>`;}
  $('#analyzeArea').innerHTML=`<div class="chartwrap"><h3>Breakdown of "${esc(d.col)}" — top ${fmt(d.shown_distinct)} by count (edit/del act on the whole database)</h3>${bars}</div>`;
  $('#analyzeArea').querySelectorAll('[data-edit]').forEach(b=>b.onclick=()=>editCat(+b.dataset.edit));
  $('#analyzeArea').querySelectorAll('[data-del]').forEach(b=>b.onclick=()=>delCat(+b.dataset.del));
}
async function editCat(i){
  const r=adata.rows[i], blank=r.value==='';
  const nv=prompt(`Rename "${blank?'(blank)':r.value}" for all ${fmt(r.count)} rows in ${st.site}.\nColumn: ${adata.col}\n(If it matches another value, they merge.)\n\nNew value:`, blank?'':r.value);
  if(nv===null)return;
  const d=await (await fetch('/api/recategorize',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({site:st.site,col:adata.col,old_value:r.value,new_value:nv,is_blank:blank})})).json();
  toast(`updated ${fmt(d.updated)} rows`); runAnalyze();
}
async function delCat(i){
  const r=adata.rows[i], blank=r.value==='';
  if(!confirm(`Delete all ${fmt(r.count)} rows where ${adata.col} = "${blank?'(blank)':r.value}" from ${st.site}?\nThis cannot be undone.`))return;
  const d=await (await fetch('/api/delete_category',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({site:st.site,col:adata.col,value:r.value,is_blank:blank})})).json();
  toast(`deleted ${fmt(d.deleted)} rows`); runAnalyze();
}
loadSites();
</script>
</body>
</html>"""
